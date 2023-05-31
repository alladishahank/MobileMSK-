import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
no_rows = sheet.max_row
for i in range(2,no_rows + 1):
    cell = sheet.cell(i,3)
    newCell = cell.value * 0.9
    newColumn = sheet.cell(i,4)
    newColumn.value = newCell

val = Reference(sheet, 2, no_rows,4,4)
chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart, 'e2')

wb.save('transaction2.xlsx')
